from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import sys # used to emulate infinity
import numpy as np

adj_matrix = list()
nodes_list = list()
algorithm_output = list()

print("Per creare il file contentente la matrice seguire il file \"prova.xlsx\"")

nodes_number = int(input("Inserisci il numero di vertici presenti nella matrice\n"))
filename = input("Inserisci il nome del file da cui prendere la matrice di adiacenza (estensione inclusa)\n")

# load excel file
wb = load_workbook(filename)

# load excel sheet
adj_ws = wb.active

# access nodes list
for i in range(2, nodes_number + 2):
    char = get_column_letter(i)
    nodes_list.append(adj_ws[char + "1"].value)

# access adj matrix table
for i in range(2, nodes_number + 2):
    matrix_row = list()
    for j in range(2, nodes_number + 2):    
        char = get_column_letter(j)
        matrix_row.append(adj_ws[char + str(i)].value)
    adj_matrix.append(matrix_row)

algorithm_output.append(["   "] + nodes_list + ["w(t)"])

# **** algorithm ****
w_tree = 0
parents_list = [None] * len(adj_matrix)
cost_list = [sys.maxsize] * len(adj_matrix)
visited = [False] * len(adj_matrix)
root = 0

cost_list[root] = 0     
parents_list[root] = nodes_list[root]

# [:] idk why this works but ok (don't remove)

algorithm_output.append(["è visitato"] + visited[:])
algorithm_output.append(["p(v)"] + parents_list[:])

# remove enormous value and substitute with "inf"
cost_list_str = cost_list[:]

for i in range(len(cost_list)):
    if (cost_list[i] == sys.maxsize):
        cost_list_str[i] = "inf"

algorithm_output.append(["c(v)"] + cost_list_str[:])

while(False in visited):

    # find minimum cost edge
    min_index = -1
    min_c = sys.maxsize
    for i in range(0, len(cost_list)):
        if(not visited[i]):
            if(cost_list[i] < min_c):
                min_c = cost_list[i]
                min_index = i

    visited[min_index] = True



    for i in range(0, len(cost_list)):
        if(adj_matrix[min_index][i] < cost_list[i] and adj_matrix[min_index][i] != 0 and not visited[i]):
            parents_list[i] = nodes_list[min_index]
            cost_list[i] = adj_matrix[min_index][i]
            w_tree += cost_list[i]

    visited_str = visited + [w_tree]

    algorithm_output.append(["è visitato"] + visited_str[:])
    algorithm_output.append(["p(v)"] + parents_list[:])

    # remove enormous value and substitute with "inf"
    cost_list_str = cost_list[:]

    for i in range(len(cost_list)):
        if (cost_list[i] == sys.maxsize):
            cost_list_str[i] = "inf"
    
    algorithm_output.append(["c(v)"] + cost_list_str[:])

if("Result" in wb.sheetnames):
    del wb["Result"]
ws_result = wb.create_sheet("Result")

for i in algorithm_output:
    print(i)

for i in algorithm_output:
    ws_result.append(i)

wb.save(filename)